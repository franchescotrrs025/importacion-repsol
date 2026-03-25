import streamlit as st
import pandas as pd
import re, io
from difflib import SequenceMatcher
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Importación Guardias Repsol", page_icon="🛢️", layout="wide")

st.markdown("""
<style>
.main-header{background:linear-gradient(90deg,#1F4E79,#2E75B6);padding:1.2rem 2rem;border-radius:10px;color:white;margin-bottom:1.5rem}
.main-header h1{margin:0;font-size:1.8rem}
.main-header p{margin:0.3rem 0 0 0;opacity:.85;font-size:.95rem}
.metric-card{background:#f0f4f8;border-left:4px solid #2E75B6;padding:.8rem 1rem;border-radius:6px;margin-bottom:.5rem}
.metric-card .val{font-size:2rem;font-weight:bold;color:#1F4E79}
.metric-card .lbl{font-size:.8rem;color:#555}
.warn-box{background:#fff3cd;border-left:4px solid #ffc107;padding:.7rem 1rem;border-radius:6px;margin-top:.5rem}
.ok-box{background:#d4edda;border-left:4px solid #28a745;padding:.7rem 1rem;border-radius:6px;margin-top:.5rem}
.step-badge{background:#1F4E79;color:white;border-radius:50%;padding:.2rem .6rem;font-weight:bold;margin-right:.5rem;font-size:.9rem}
.legend-box{background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:.8rem 1rem;font-size:.85rem}
.legend-box table{width:100%;border-collapse:collapse}
.legend-box td{padding:.25rem .5rem}
.badge-p{background:#d4edda;color:#155724;border-radius:4px;padding:.1rem .4rem;font-weight:bold}
.badge-fer{background:#cce5ff;color:#004085;border-radius:4px;padding:.1rem .4rem;font-weight:bold}
.badge-a{background:#fff3cd;color:#856404;border-radius:4px;padding:.1rem .4rem;font-weight:bold}
.badge-d{background:#f8d7da;color:#721c24;border-radius:4px;padding:.1rem .4rem;font-weight:bold}
.badge-l{background:#e2e3e5;color:#383d41;border-radius:4px;padding:.1rem .4rem;font-weight:bold}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🛢️ Importación de Guardias — Repsol Nuevo Lote 57</h1>
    <p>Sube los archivos de guardias y activos para generar el Excel de importación al sistema.</p>
</div>
""", unsafe_allow_html=True)

# ── CONSTANTES ────────────────────────────────────────
CODIGOS = {
    "CHEF":523,"COCINERO":523,"PANADERO PASTELERO":628,"PAN/PASTELERO":628,
    "PANADEROS":628,"AYUDANTE DE COCINA":523,"AYD.COCINA":523,
    "MOZO":629,"AZAFATA":629,"VAJILLERO":528,
    "AUXILIAR DE HOTELERIA":524,"HOTELERO":524,"LIDER DE HOTELERIA":524,
    "LIDER DE GRUPO":524,"LAVANDERO":525,
    "ALMACENERO":630,"AUXILIAR DE ALMACEN":630,"AUX. ALMACEN":630,
    "SUPERVISOR DE SEGURIDAD":528,"SUPERVISOR DE CALIDAD":528,
    "SUP.CALIDAD":528,"SUP.HSE":528,"JEFE CONTRATO":528,
    "JEFE DE CAMPAMENTO - TRAINING":528,"AUXILIAR ADMINISTRATIVO":528,
    "JEFE DE CONTRATO":528,"AUXILIAR DE ADMINISTRACION":528,
    "NUTRICIONISTA":528,"INSTRUCTORA":528,
}
MESES = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
         'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
ANIO  = datetime.now().year
STOP  = {"PERSONAL EN CAMPO","PERSONAL DESCANSO","TOTAL LOGISTICA",
         "TOTAL CATERING","TOTAL HOTELERIA","TOTALES"}

# ── LEYENDA DE ASISTENCIA ─────────────────────────────
# P   = Presente              → código de dirección (trabajó)
# FER = Feriado Trabajado     → código de dirección (trabajó)
# A   = Vacaciones            → -1
# D   = Día de Descanso       → -1
# L   = Licencia sin Goce     → -1
# Número (horas: 12, 13...)   → código de dirección (trabajó)
# Vacío / NaN                 → -1

CODIGOS_TRABAJO  = {"P", "FER"}               # Asistencia → trabajó
CODIGOS_DESCANSO = {"A", "D", "L", "DES", "VACACIONES", "VAC"}  # Ausencia → -1

# ── HELPERS ───────────────────────────────────────────
def parsear_rango(texto):
    t = str(texto).strip().lower()
    m = re.match(r'(\d+)\s+al\s+(\d+)\s+(\w{3})', t)
    if not m: return []
    d_ini, d_fin, mes_str = int(m.group(1)), int(m.group(2)), m.group(3)
    mes = MESES.get(mes_str)
    if not mes: return []
    fecha_fin = date(ANIO, mes, d_fin)
    mes_ini = (mes-1) if d_fin < d_ini else mes
    anio_ini = ANIO if mes_ini > 0 else ANIO-1
    if mes_ini == 0: mes_ini = 12
    fecha_ini = date(anio_ini, mes_ini, d_ini)
    dias = []
    f = fecha_ini
    while f <= fecha_fin:
        dias.append(f); f += timedelta(days=1)
    return dias

def a_codigo(val, cargo):
    """
    Interpreta el valor de celda según la nueva leyenda de asistencia:
      P / FER / número de horas  → código de dirección del cargo (trabajó)
      A / D / L                  → -1  (ausencia o descanso)
      vacío / NaN                → -1
    Mantiene compatibilidad con el formato anterior (1 → trabajó).
    """
    if pd.isna(val) or val is None:
        return -1
    s = str(val).strip().upper()
    if s in ["", "NAN", "0", "0.0"]:
        return -1
    # Vacaciones explícitas (texto largo)
    if "VACACION" in s or s.startswith("SSS"):
        return -1
    # Nuevos códigos: trabajo
    if s in CODIGOS_TRABAJO:
        return CODIGOS.get(cargo.upper(), 528)
    # Nuevos códigos: descanso/ausencia
    if s in CODIGOS_DESCANSO:
        return -1
    # Legado: "1" = trabajó
    if s in ["1", "1.0"]:
        return CODIGOS.get(cargo.upper(), 528)
    # Número de horas trabajadas (ej. 12, 13, 8) → trabajó
    try:
        horas = float(s)
        if horas > 0:
            return CODIGOS.get(cargo.upper(), 528)
        return -1
    except ValueError:
        pass
    return -1

def clasificar_celda(val, cargo):
    """Devuelve (codigo, etiqueta_visual) para el preview."""
    cod = a_codigo(val, cargo)
    if pd.isna(val) or val is None or str(val).strip().upper() in ["", "NAN", "0", "0.0"]:
        return cod, "—"
    s = str(val).strip().upper()
    if s == "P":    return cod, "P"
    if s == "FER":  return cod, "FER"
    if s == "A":    return cod, "A"
    if s == "D":    return cod, "D"
    if s == "L":    return cod, "L"
    try: float(s); return cod, s  # horas
    except: pass
    return cod, s

def cargar_activos(bytes_io):
    df = pd.read_excel(bytes_io)
    col_nombre = next((c for c in df.columns if "NOMBRE" in c.upper()), None)
    col_dni = next((c for c in df.columns if "NRO" in c.upper() and "DOCUMENTO" in c.upper()), None)
    if col_dni is None:
        col_dni = next((c for c in df.columns if "DOCUMENTO" in c.upper() and "TIPO" not in c.upper()), None)
    if col_dni is None:
        col_dni = next((c for c in df.columns if "DNI" in c.upper()), None)
    if not col_nombre or not col_dni:
        raise ValueError(f"No se encontraron columnas de NOMBRE y DOCUMENTO. Columnas: {list(df.columns)}")
    df["_nc"] = (df[col_nombre].astype(str).str.upper().str.strip()
                 .str.replace(r"[,.]", "", regex=True).str.replace(r"\s+", " ", regex=True))
    df["_dni"] = df[col_dni].astype(str).str.strip()
    import unicodedata
    def norm_col(t):
        t2 = re.sub(r"[,.]", "", str(t).upper().strip())
        t2 = re.sub(r"\s+", " ", t2).strip()
        t3 = unicodedata.normalize("NFKD", t2).encode("ascii","ignore").decode("ascii")
        return t3.strip()
    df["_nc_norm"] = df["_nc"].apply(norm_col)
    return df

def normalizar(texto):
    import unicodedata
    t = re.sub(r"[,.]", "", str(texto).upper().strip())
    t = re.sub(r"\s+", " ", t).strip()
    t_norm = unicodedata.normalize("NFKD", t)
    t_ascii = t_norm.encode("ascii", "ignore").decode("ascii")
    return t_ascii.strip()

def buscar_dni(nombre, activos_df):
    n      = normalizar(nombre)
    n_orig = re.sub(r"[,.]", "", str(nombre).upper().strip())
    n_orig = re.sub(r"\s+", " ", n_orig).strip()
    if not n or n in ["NAN", ""]: return "NO ENCONTRADO"
    m = activos_df[activos_df["_nc"] == n_orig]
    if not m.empty: return m.iloc[0]["_dni"]
    m = activos_df[activos_df["_nc_norm"] == n]
    if not m.empty: return m.iloc[0]["_dni"]
    palabras      = set(n.split())
    palabras_orig = set(n_orig.split())
    for _, row in activos_df.iterrows():
        pa = set(row["_nc"].split())
        pa_n = set(row["_nc_norm"].split())
        if len(palabras_orig) >= 2 and palabras_orig == pa: return row["_dni"]
        if len(palabras) >= 2     and palabras == pa_n:     return row["_dni"]
    for _, row in activos_df.iterrows():
        pa_n = set(row["_nc_norm"].split())
        if len(palabras) >= 2 and palabras.issubset(pa_n):
            return row["_dni"]
    if len(palabras) >= 2:
        for _, row in activos_df.iterrows():
            pa_n = set(row["_nc_norm"].split())
            if len(palabras & pa_n) >= min(3, len(palabras)):
                return row["_dni"]
    n_sin_esp = n.replace(" ", "")
    for _, row in activos_df.iterrows():
        if row["_nc_norm"].replace(" ", "") == n_sin_esp:
            return row["_dni"]
    mejor_score = 0; mejor_dni = "NO ENCONTRADO"
    for _, row in activos_df.iterrows():
        score = SequenceMatcher(None, n_sin_esp, row["_nc_norm"].replace(" ","")).ratio()
        if score > mejor_score:
            mejor_score = score; mejor_dni = row["_dni"]
    if mejor_score >= 0.88:
        return mejor_dni
    return "NO ENCONTRADO"

def es_nombre_valido(texto):
    if not texto or texto in ["NAN","","0","0.0"]: return False
    if texto.startswith("SSS"): return False
    if texto == "HOTELERIA": return False
    try: float(texto); return False
    except: pass
    return True

def parsear_hoja_unica(df, activos_df):
    resultados = []; i = 0
    while i < len(df):
        row = df.iloc[i]
        c0 = str(row.iloc[0]).strip().upper() if pd.notna(row.iloc[0]) else ""
        c1 = str(row.iloc[1]).strip().upper() if pd.notna(row.iloc[1]) else ""

        if ("GUARDI" in c0 or "GUADI" in c0) and c1 == "CARGO":
            col_fechas = {}
            for ci, val in row.items():
                if isinstance(val, (datetime, pd.Timestamp)):
                    col_fechas[ci] = [pd.Timestamp(val).date()]
                elif pd.notna(val) and isinstance(val, str):
                    dias = parsear_rango(val)
                    if dias: col_fechas[ci] = dias
            if not col_fechas: i += 1; continue

            i += 1; guardia = None
            while i < len(df):
                fila = df.iloc[i]
                g     = str(fila.iloc[0]).strip() if pd.notna(fila.iloc[0]) else ""
                cargo = str(fila.iloc[1]).strip().upper() if pd.notna(fila.iloc[1]) else ""
                nom_h = str(fila.iloc[2]).strip().upper() if pd.notna(fila.iloc[2]) else ""

                if nom_h in STOP or cargo in STOP: break
                if ("GUARDI" in g.upper() or "GUADI" in g.upper()) and cargo == "CARGO": break

                if g and g.upper() not in ["NAN",""]: guardia = g
                if not cargo or cargo in ["NAN","APELLIDOS Y NOMBRES"]: i += 1; continue
                if not es_nombre_valido(nom_h): i += 1; continue

                dias = {}
                for ci, fechas_lista in col_fechas.items():
                    val = fila.iloc[ci] if ci < len(fila) else None
                    for f in fechas_lista: dias[f] = val

                resultados.append({
                    "guardia": guardia, "cargo": cargo, "nombre": nom_h,
                    "dni": buscar_dni(nom_h, activos_df),
                    "dias": dias,
                })
                i += 1
            continue
        i += 1
    return resultados

def parsear_hoja_formato_nuevo(df, activos_df):
    """
    Parsea el nuevo formato con leyenda P/A/D/FER/L en la cabecera.
    Detecta: fila de leyenda (P=Presente, A=Vacaciones...), luego la fila
    de fechas (row 15-17 en la imagen) y los datos a partir de ahí.
    Columnas: Matrícula | Nombre | fecha1_col1 | fecha1_col2 | fecha2_col1 ...
    Cada fecha puede tener sub-columnas (H, T, indicador de asistencia).
    Se toma la sub-columna que contenga P/A/D/FER/L o número de horas.
    """
    resultados = []

    # 1. Buscar fila de encabezado de fechas (contiene objetos datetime o strings de fecha)
    fila_fechas_idx = None
    for ri in range(len(df)):
        row = df.iloc[ri]
        fechas_en_fila = 0
        for val in row:
            if isinstance(val, (datetime, pd.Timestamp)):
                fechas_en_fila += 1
        if fechas_en_fila >= 3:
            fila_fechas_idx = ri
            break

    if fila_fechas_idx is None:
        return []  # No se encontró la fila de fechas

    # 2. Construir mapa columna → fecha
    col_a_fecha = {}
    fila_h = df.iloc[fila_fechas_idx]
    for ci, val in enumerate(fila_h):
        if isinstance(val, (datetime, pd.Timestamp)):
            col_a_fecha[ci] = pd.Timestamp(val).date()

    if not col_a_fecha:
        return []

    # 3. Buscar fila de inicio de datos (contiene "Matricula"/"Nombre" o la primera matrícula numérica)
    datos_inicio = fila_fechas_idx + 1
    # Saltar sub-encabezados (H, T, 1, -)
    for ri in range(fila_fechas_idx + 1, min(fila_fechas_idx + 5, len(df))):
        row = df.iloc[ri]
        c0 = str(row.iloc[0]).strip().upper() if pd.notna(row.iloc[0]) else ""
        c1 = str(row.iloc[1]).strip().upper() if pd.notna(row.iloc[1]) else ""
        if "MATRICU" in c0 or "NOMBRE" in c1:
            datos_inicio = ri + 1
            break
        # Si la primera celda parece un número de matrícula → datos empiezan aquí
        try:
            int(str(row.iloc[0]).strip())
            datos_inicio = ri
            break
        except:
            pass

    # 4. Recorrer filas de datos
    # Determinar cargo desde contexto (puede estar en una fila anterior de sección)
    cargo_actual = "HOTELERO"  # default

    for ri in range(datos_inicio, len(df)):
        row = df.iloc[ri]
        c0 = str(row.iloc[0]).strip().upper() if pd.notna(row.iloc[0]) else ""
        c1 = str(row.iloc[1]).strip().upper() if pd.notna(row.iloc[1]) else ""

        # Detectar cambio de cargo/sección
        if c0 in STOP or c1 in STOP:
            continue
        if c0 in CODIGOS:
            cargo_actual = c0
            continue

        # Verificar que c0 parece matrícula (número) y c1 parece nombre
        try:
            int(c0.replace(" ",""))
            matricula = c0
        except:
            # Si no es número, puede ser un nombre en c0
            if not es_nombre_valido(c1):
                continue
            matricula = ""

        nombre = c1
        if not es_nombre_valido(nombre):
            nombre = c0
            if not es_nombre_valido(nombre):
                continue

        # Recopilar días: para cada fecha, buscar el valor de asistencia en esa columna
        dias = {}
        for ci, fecha in col_a_fecha.items():
            val = row.iloc[ci] if ci < len(row) else None
            dias[fecha] = val

        resultados.append({
            "guardia": "GUARDIA",
            "cargo": cargo_actual,
            "nombre": nombre,
            "dni": buscar_dni(nombre, activos_df),
            "dias": dias,
        })

    return resultados

def procesar(guardias_bytes, activos_bytes):
    activos_df = cargar_activos(activos_bytes)
    xl = pd.ExcelFile(guardias_bytes)
    hojas = xl.sheet_names

    todas = []; fechas_set = set()
    for hoja in hojas:
        if hoja.upper() in ["ESTRUCTURA"]: continue
        df = pd.read_excel(guardias_bytes, sheet_name=hoja, header=None)

        # Intentar formato clásico (encabezado GUARDIAS / CARGO)
        filas = parsear_hoja_unica(df, activos_df)

        # Si no encontró nada, intentar nuevo formato (Matricula / Nombre / fechas)
        if not filas:
            filas = parsear_hoja_formato_nuevo(df, activos_df)

        for f in filas: fechas_set.update(f["dias"].keys())
        todas.extend(filas)

    return todas, sorted(fechas_set)

def generar_excel(todas, fechas_ord):
    wb = Workbook(); ws = wb.active; ws.title = "Importacion"
    hf  = PatternFill("solid", start_color="1F4E79")
    sf  = PatternFill("solid", start_color="2E75B6")
    df2 = PatternFill("solid", start_color="D9D9D9")
    af  = PatternFill("solid", start_color="E2EFDA")
    ef  = PatternFill("solid", start_color="FFDCE0")
    yf  = PatternFill("solid", start_color="FFF3CD")  # Vacaciones
    rf2 = PatternFill("solid", start_color="F8D7DA")  # Descanso
    bf  = PatternFill("solid", start_color="CCE5FF")  # Feriado trabajado
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    dfont = Font(name="Arial", size=9)
    bfont = Font(bold=True, name="Arial", size=9)
    gfont = Font(bold=True, name="Arial", size=9, color="375623")
    rfont = Font(bold=True, name="Arial", size=9, color="C00000")
    cen = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")
    th  = Side(border_style="thin", color="BFBFBF")
    brd = Border(left=th, right=th, top=th, bottom=th)

    lc = get_column_letter(4+len(fechas_ord))
    ws.merge_cells(f"A1:{lc}1")
    ws["A1"] = "IMPORTACIÓN GUARDIAS REPSOL — NUEVO LOTE 57"
    ws["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws["A1"].fill = hf; ws["A1"].alignment = cen; ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{lc}2")
    rng = f"Período: {fechas_ord[0].strftime('%d/%m/%Y')} al {fechas_ord[-1].strftime('%d/%m/%Y')}" if fechas_ord else ""
    ws["A2"] = rng; ws["A2"].font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ws["A2"].fill = sf; ws["A2"].alignment = cen; ws.row_dimensions[2].height = 15

    # Fila de leyenda
    leyenda_cols = ["P = Presente", "FER = Feriado Trabajado", "A = Vacaciones", "D = Descanso", "L = Licencia s/Goce"]
    for idx, txt in enumerate(leyenda_cols):
        c = ws.cell(row=3, column=1+idx, value=txt)
        colors_leg = ["E2EFDA","CCE5FF","FFF3CD","F8D7DA","E2E3E5"]
        c.fill = PatternFill("solid", start_color=colors_leg[idx])
        c.font = Font(bold=True, name="Arial", size=8)
        c.alignment = cen; c.border = brd
    # Rellenar resto de leyenda con vacío
    for idx in range(5, 4+len(fechas_ord)):
        ws.cell(row=3, column=1+idx).border = brd
    ws.row_dimensions[3].height = 14

    hdrs = ["APELLIDOS Y NOMBRES","DNI","CARGO","DIRECCIÓN ID"] + [f.strftime("%d/%m/%Y") for f in fechas_ord]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = hfont; c.fill = sf; c.alignment = cen; c.border = brd
    ws.row_dimensions[4].height = 18
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 26; ws.column_dimensions["D"].width = 14
    for i in range(len(fechas_ord)):
        ws.column_dimensions[get_column_letter(5+i)].width = 10

    re2 = 5
    for e in todas:
        cargo = e["cargo"]; dni = e["dni"]
        cod = CODIGOS.get(cargo.upper(), 528)
        vals = [e["nombre"], dni, cargo, cod] + [a_codigo(e["dias"].get(f), cargo) for f in fechas_ord]
        raw_vals = [e["dias"].get(f) for f in fechas_ord]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=re2, column=ci, value=v); c.border = brd; c.font = dfont
            if ci <= 4:
                c.alignment = lft
                if ci == 1: c.font = bfont
                if ci == 2 and dni == "NO ENCONTRADO": c.fill = ef; c.font = rfont
            else:
                c.alignment = cen
                raw = raw_vals[ci-5] if (ci-5) < len(raw_vals) else None
                raw_s = str(raw).strip().upper() if raw and not pd.isna(raw) else ""
                if v == -1:
                    if raw_s == "A":   c.fill = yf  # Vacaciones → amarillo
                    elif raw_s == "D": c.fill = rf2  # Descanso → rojo claro
                    elif raw_s == "L": c.fill = PatternFill("solid", start_color="E2E3E5")
                    else:              c.fill = df2
                    c.value = -1
                else:
                    if raw_s == "FER": c.fill = bf; c.font = Font(bold=True, name="Arial", size=9, color="004085")
                    else:              c.fill = af; c.font = gfont
        re2 += 1

    ws.cell(row=re2, column=1, value=f"TOTAL: {re2-5} registros").font = bfont
    ws.freeze_panes = "E5"

    # Hoja de leyenda completa
    ws3 = wb.create_sheet("Leyenda Asistencia")
    ley_data = [
        ("CÓDIGO","DESCRIPCIÓN","RESULTADO EN IMPORTACIÓN","COLOR"),
        ("P","Presente","Código de dirección del área","Verde"),
        ("FER","Feriado Trabajado","Código de dirección del área","Azul claro"),
        ("A","Vacaciones","-1","Amarillo"),
        ("D","Día de Descanso","-1","Rojo claro"),
        ("L","Licencia sin Goce","-1","Gris"),
        ("(vacío)","Sin registro","-1","Gris"),
        ("12/13/...","Horas trabajadas","Código de dirección del área","Verde"),
    ]
    ley_fills = [sf, af, bf, yf, rf2,
                 PatternFill("solid", start_color="E2E3E5"),
                 df2, af]
    for ri, (row_data, fill) in enumerate(zip(ley_data, ley_fills), 1):
        for ci, val in enumerate(row_data, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.border = brd; c.alignment = cen
            if ri == 1: c.font = hfont; c.fill = sf
            else:       c.font = dfont; c.fill = fill
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 26
    ws3.column_dimensions["C"].width = 34
    ws3.column_dimensions["D"].width = 14

    # Hoja de códigos de dirección
    ws2 = wb.create_sheet("Códigos de Dirección")
    for ci, h in enumerate(["CARGO","CÓDIGO","ÁREA"],1):
        c = ws2.cell(row=1,column=ci,value=h); c.font=hfont; c.fill=sf; c.alignment=cen; c.border=brd
    ws2.column_dimensions["A"].width=35; ws2.column_dimensions["B"].width=12; ws2.column_dimensions["C"].width=30
    adesc = {523:"Cocina T.R.2",524:"Hotelería",525:"Lavandería",528:"Vajilla/Admin",629:"Mozo",630:"Almacén",628:"Panadería T.R.2"}
    for i,(k,v) in enumerate(CODIGOS.items(),2):
        ws2.cell(row=i,column=1,value=k).border=brd
        c=ws2.cell(row=i,column=2,value=v); c.border=brd; c.alignment=cen
        ws2.cell(row=i,column=3,value=adesc.get(v,"")).border=brd

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── HISTORIAL ─────────────────────────────────────────
if "historial" not in st.session_state:
    st.session_state.historial = []

# ── SIDEBAR ───────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📋 Historial de importaciones")
    if st.session_state.historial:
        for h in reversed(st.session_state.historial[-10:]):
            st.markdown(f"""
            <div style='background:#f8f9fa;border-radius:6px;padding:.5rem .7rem;margin-bottom:.4rem;font-size:.82rem;border-left:3px solid #2E75B6'>
            🕐 <b>{h['fecha']}</b><br>
            👥 {h['registros']} registros · 📅 {h['dias']} días<br>
            ⚠️ {h['sin_dni']} sin DNI
            </div>""", unsafe_allow_html=True)
    else:
        st.info("Sin importaciones aún.")
    st.markdown("---")
    st.markdown("### 🗂️ Leyenda de asistencia")
    st.markdown("""
    <div class="legend-box">
    <table>
    <tr><td><span class="badge-p">P</span></td><td>Presente → <b>código área</b></td></tr>
    <tr><td><span class="badge-fer">FER</span></td><td>Feriado trabajado → <b>código área</b></td></tr>
    <tr><td><span class="badge-a">A</span></td><td>Vacaciones → <b>-1</b></td></tr>
    <tr><td><span class="badge-d">D</span></td><td>Día de descanso → <b>-1</b></td></tr>
    <tr><td><span class="badge-l">L</span></td><td>Licencia sin goce → <b>-1</b></td></tr>
    <tr><td><b>12/13</b></td><td>Horas → <b>código área</b></td></tr>
    </table>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 🗂️ Códigos de dirección")
    adesc = {523:"Cocina",524:"Hotelería",525:"Lavandería",528:"Vajilla/Admin",629:"Mozo",630:"Almacén",628:"Panadería"}
    codigos_vistos = set()
    for cargo, cod in CODIGOS.items():
        if cod not in codigos_vistos:
            st.markdown(f"<small>`{cod}` — {adesc.get(cod,'')} </small>", unsafe_allow_html=True)
            codigos_vistos.add(cod)

# ── MAIN ──────────────────────────────────────────────
st.markdown('<span class="step-badge">1</span> **Sube los archivos Excel**', unsafe_allow_html=True)

# Mostrar leyenda inline
st.markdown("""
<div class="legend-box" style="margin-bottom:1rem">
<b>Leyenda de asistencia (nuevo formato):</b> &nbsp;
<span class="badge-p">P</span> Presente &nbsp;
<span class="badge-fer">FER</span> Feriado Trabajado &nbsp;
<span class="badge-a">A</span> Vacaciones &nbsp;
<span class="badge-d">D</span> Descanso &nbsp;
<span class="badge-l">L</span> Licencia &nbsp;
<b>12/13</b> Horas trabajadas
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    f_guardias = st.file_uploader("📁 Archivo de Guardias Repsol", type=["xlsx"],
        help="El archivo con las guardias (formato con P/A/D/FER/L o formato clásico con 1)")
with col2:
    f_activos = st.file_uploader("📁 Archivo de Activos (con DNIs)", type=["xlsx"],
        help="El Excel con la lista de trabajadores activos y sus DNIs")

st.markdown("---")

if f_guardias and f_activos:
    st.markdown('<span class="step-badge">2</span> **Procesando archivos...**', unsafe_allow_html=True)
    with st.spinner("Leyendo guardias y buscando DNIs..."):
        try:
            guardias_bytes = io.BytesIO(f_guardias.read())
            activos_bytes  = io.BytesIO(f_activos.read())
            todas, fechas_ord = procesar(guardias_bytes, activos_bytes)
        except Exception as ex:
            st.error(f"❌ Error: {ex}")
            try:
                guardias_bytes.seek(0)
                xl = pd.ExcelFile(guardias_bytes)
                st.warning(f"📋 Hojas en el archivo de guardias: {xl.sheet_names}")
            except: pass
            st.stop()

    if not todas:
        st.warning("No se encontraron registros. Verificá que el archivo tenga el formato correcto.")
        st.stop()

    sin_dni = [e for e in todas if e["dni"] == "NO ENCONTRADO"]

    st.markdown('<span class="step-badge">3</span> **Resumen**', unsafe_allow_html=True)
    m1,m2,m3,m4 = st.columns(4)
    with m1:
        st.markdown(f'<div class="metric-card"><div class="val">{len(todas)}</div><div class="lbl">👥 Registros</div></div>', unsafe_allow_html=True)
    with m2:
        st.markdown(f'<div class="metric-card"><div class="val">{len(fechas_ord)}</div><div class="lbl">📅 Días</div></div>', unsafe_allow_html=True)
    with m3:
        color = "#C00000" if sin_dni else "#28a745"
        st.markdown(f'<div class="metric-card"><div class="val" style="color:{color}">{len(sin_dni)}</div><div class="lbl">⚠️ Sin DNI</div></div>', unsafe_allow_html=True)
    with m4:
        rango = f"{fechas_ord[0].strftime('%d/%m')} → {fechas_ord[-1].strftime('%d/%m/%Y')}" if fechas_ord else "—"
        st.markdown(f'<div class="metric-card"><div class="val" style="font-size:1rem;padding-top:.5rem">{rango}</div><div class="lbl">📆 Período</div></div>', unsafe_allow_html=True)

    # Conteo por tipo de asistencia
    conteo_p = conteo_fer = conteo_a = conteo_d = conteo_l = 0
    for e in todas:
        for f in fechas_ord:
            v = str(e["dias"].get(f, "")).strip().upper()
            if v == "P": conteo_p += 1
            elif v == "FER": conteo_fer += 1
            elif v == "A": conteo_a += 1
            elif v == "D": conteo_d += 1
            elif v == "L": conteo_l += 1

    st.markdown(f"""
    <div style='display:flex;gap:.5rem;flex-wrap:wrap;margin-bottom:.5rem'>
    <span class="badge-p">P {conteo_p}</span>
    <span class="badge-fer">FER {conteo_fer}</span>
    <span class="badge-a">A {conteo_a}</span>
    <span class="badge-d">D {conteo_d}</span>
    <span class="badge-l">L {conteo_l}</span>
    </div>
    """, unsafe_allow_html=True)

    if sin_dni:
        nombres_sin = [e["nombre"] for e in sin_dni]
        st.markdown(f'<div class="warn-box">⚠️ <b>{len(nombres_sin)} sin DNI:</b><br>{"<br>".join(f"• {n}" for n in nombres_sin)}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="ok-box">✅ <b>Todos los DNIs encontrados.</b></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<span class="step-badge">4</span> **Vista previa**', unsafe_allow_html=True)

    preview = []
    for e in todas:
        dias_trabajo = sum(1 for f in fechas_ord if a_codigo(e["dias"].get(f), e["cargo"]) != -1)
        dias_vac     = sum(1 for f in fechas_ord if str(e["dias"].get(f,"")).strip().upper() == "A")
        dias_desc    = sum(1 for f in fechas_ord if str(e["dias"].get(f,"")).strip().upper() == "D")
        preview.append({
            "Nombre": e["nombre"], "DNI": e["dni"], "Cargo": e["cargo"],
            "Cód.Dir": CODIGOS.get(e["cargo"].upper(), 528),
            "✅ Trabajo": dias_trabajo,
            "🏖️ Vacac.": dias_vac,
            "😴 Descanso": dias_desc,
        })

    df_prev = pd.DataFrame(preview)
    fc1, fc2 = st.columns([2,1])
    with fc2:
        solo_sin = st.checkbox("Solo sin DNI")
    df_show = df_prev[df_prev["DNI"]=="NO ENCONTRADO"] if solo_sin else df_prev

    def hl(row): return ["background-color:#FFDCE0"]*len(row) if row["DNI"]=="NO ENCONTRADO" else [""]*len(row)
    st.dataframe(df_show.style.apply(hl, axis=1), use_container_width=True, height=350)

    st.markdown("---")
    st.markdown('<span class="step-badge">5</span> **Descargar Excel**', unsafe_allow_html=True)

    guardias_bytes.seek(0)
    excel_buf = generar_excel(todas, fechas_ord)
    fecha_hoy = datetime.now().strftime("%Y%m%d_%H%M")

    st.download_button(
        label="⬇️ Descargar Excel de Importación",
        data=excel_buf,
        file_name=f"Importacion_Repsol_{fecha_hoy}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )

    entrada = {"fecha": datetime.now().strftime("%d/%m/%Y %H:%M"),
               "registros": len(todas), "dias": len(fechas_ord), "sin_dni": len(sin_dni)}
    if not st.session_state.historial or st.session_state.historial[-1]["fecha"] != entrada["fecha"]:
        st.session_state.historial.append(entrada)
else:
    st.info("👆 Sube ambos archivos para comenzar.")
    st.markdown("""
    **¿Qué hace esta app?**
    1. Lee el archivo de **guardias** (formato nuevo P/A/D/FER/L o formato clásico)
    2. Busca el **DNI** de cada trabajador en el archivo de Activos
    3. **P / FER / horas** → reemplaza por el **código de dirección** del área
    4. **A / D / L** → marca como **-1** (ausencia/descanso)
    5. Genera el **Excel listo para importar** con colores por tipo
    """)
